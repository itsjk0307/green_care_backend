from __future__ import annotations

import calendar
import io
import uuid
from datetime import date, datetime, timedelta, timezone

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import cast, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy.types import Date as SaDate

from app.models.daily_work_plan import DailyWorkerAttendance, DailyWorkPlan
from app.models.golf_course import GolfCourse
from app.models.issue import Issue
from app.models.user import User
from app.models.work_report import WorkReport
from app.schemas.journal import (
    AttendanceItemResponse,
    AttendanceSectionResponse,
    AttendanceSummaryResponse,
    DailyJournalResponse,
    MonthlyDaySummaryResponse,
    MonthlyJournalResponse,
    WeatherInfoResponse,
    ZoneTaskJournalItemResponse,
)


async def _ensure_course_exists(db: AsyncSession, course_id: uuid.UUID) -> None:
    result = await db.execute(select(GolfCourse.id).where(GolfCourse.id == course_id))
    if result.scalar_one_or_none() is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Golf course not found.",
        )


def _parse_worker_ids(raw_ids: list[str]) -> list[uuid.UUID]:
    parsed: list[uuid.UUID] = []
    for raw in raw_ids:
        try:
            parsed.append(uuid.UUID(str(raw)))
        except ValueError:
            continue
    return parsed


async def _load_users_map(db: AsyncSession, user_ids: list[uuid.UUID]) -> dict[uuid.UUID, User]:
    if not user_ids:
        return {}
    result = await db.execute(select(User).where(User.id.in_(user_ids)))
    return {user.id: user for user in result.scalars().all()}


async def get_daily_journal(
    db: AsyncSession,
    *,
    course_id: uuid.UUID,
    journal_date: date,
) -> DailyJournalResponse:
    await _ensure_course_exists(db, course_id)

    plan_result = await db.execute(
        select(DailyWorkPlan)
        .where(
            DailyWorkPlan.course_id == course_id,
            DailyWorkPlan.plan_date == journal_date,
        )
        .options(
            selectinload(DailyWorkPlan.zone_tasks),
            selectinload(DailyWorkPlan.attendance).selectinload(DailyWorkerAttendance.worker),
        )
    )
    plan = plan_result.scalar_one_or_none()

    if plan is None:
        weather_info = WeatherInfoResponse(plan_exists=False)
        attendance_section = AttendanceSectionResponse()
        zone_tasks: list[ZoneTaskJournalItemResponse] = []
    else:
        weather_info = WeatherInfoResponse(
            weather=plan.weather,
            temperature_min=plan.temperature_min,
            temperature_max=plan.temperature_max,
            rainfall_mm=plan.rainfall_mm,
            special_notes=plan.special_notes,
            status=plan.status,
            total_workers=plan.total_workers,
            plan_exists=True,
        )

        attendance_items: list[AttendanceItemResponse] = []
        total_present = 0
        total_absent = 0
        total_overtime = 0
        for record in plan.attendance:
            worker_name = record.worker.name if record.worker else "Unknown"
            attendance_items.append(
                AttendanceItemResponse(
                    worker_name=worker_name,
                    status=record.status,
                    start_time=record.start_time,
                    end_time=record.end_time,
                    working_hours=record.working_hours,
                )
            )
            if record.status == "present":
                total_present += 1
            elif record.status == "absent":
                total_absent += 1
            elif record.status == "overtime":
                total_overtime += 1

        attendance_section = AttendanceSectionResponse(
            items=attendance_items,
            summary=AttendanceSummaryResponse(
                total_present=total_present,
                total_absent=total_absent,
                total_overtime=total_overtime,
            ),
        )

        all_worker_ids: set[uuid.UUID] = set()
        for task in plan.zone_tasks:
            all_worker_ids.update(_parse_worker_ids(task.assigned_worker_ids))
        users_map = await _load_users_map(db, list(all_worker_ids))

        zone_tasks = []
        for task in plan.zone_tasks:
            names: list[str] = []
            for worker_id in _parse_worker_ids(task.assigned_worker_ids):
                user = users_map.get(worker_id)
                if user:
                    names.append(user.name)
            zone_tasks.append(
                ZoneTaskJournalItemResponse(
                    zone=task.zone,
                    task_types=task.task_types,
                    mowing_height_mm=task.mowing_height_mm,
                    status=task.status,
                    completed_at=task.completed_at,
                    assigned_worker_names=names,
                )
            )

    photo_count = int(
        (
            await db.execute(
                select(func.count(WorkReport.id)).where(
                    WorkReport.course_id == course_id,
                    cast(WorkReport.created_at, SaDate) == journal_date,
                )
            )
        ).scalar_one()
    )

    issues_opened = int(
        (
            await db.execute(
                select(func.count(Issue.id)).where(
                    Issue.course_id == course_id,
                    cast(Issue.created_at, SaDate) == journal_date,
                )
            )
        ).scalar_one()
    )

    issues_resolved = int(
        (
            await db.execute(
                select(func.count(Issue.id)).where(
                    Issue.course_id == course_id,
                    Issue.resolved_at.isnot(None),
                    cast(Issue.resolved_at, SaDate) == journal_date,
                )
            )
        ).scalar_one()
    )

    return DailyJournalResponse(
        course_id=course_id,
        journal_date=journal_date,
        weather_info=weather_info,
        attendance=attendance_section,
        zone_tasks=zone_tasks,
        photo_count=photo_count,
        issues_opened=issues_opened,
        issues_resolved=issues_resolved,
    )


async def get_monthly_journal(
    db: AsyncSession,
    *,
    course_id: uuid.UUID,
    month: str,
) -> MonthlyJournalResponse:
    await _ensure_course_exists(db, course_id)

    try:
        year_str, month_str = month.split("-", 1)
        year = int(year_str)
        mon = int(month_str)
        if mon < 1 or mon > 12:
            raise ValueError("invalid month")
        start = date(year, mon, 1)
        days_in_month = calendar.monthrange(year, mon)[1]
        end = date(year, mon, days_in_month)
    except ValueError as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="month must be in YYYY-MM format.",
        ) from error

    result = await db.execute(
        select(DailyWorkPlan)
        .where(
            DailyWorkPlan.course_id == course_id,
            DailyWorkPlan.plan_date >= start,
            DailyWorkPlan.plan_date <= end,
        )
        .options(selectinload(DailyWorkPlan.zone_tasks))
    )
    plans = {plan.plan_date: plan for plan in result.scalars().all()}

    day_summaries: list[MonthlyDaySummaryResponse] = []
    for day_offset in range(days_in_month):
        current = start + timedelta(days=day_offset)
        plan = plans.get(current)
        if plan is None:
            day_summaries.append(
                MonthlyDaySummaryResponse(
                    date=current,
                    plan_exists=False,
                )
            )
            continue

        total_tasks = len(plan.zone_tasks)
        completed_tasks = sum(1 for task in plan.zone_tasks if task.status == "done")
        completion_pct = (
            round((completed_tasks / total_tasks) * 100.0, 2) if total_tasks > 0 else 0.0
        )
        day_summaries.append(
            MonthlyDaySummaryResponse(
                date=current,
                plan_exists=True,
                completion_pct=completion_pct,
                worker_count=plan.total_workers,
                weather=plan.weather,
            )
        )

    return MonthlyJournalResponse(
        course_id=course_id,
        month=month,
        days=day_summaries,
    )


def _style_header_row(worksheet, headers: list[str], fill_color: str = "1B5E20") -> None:
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_fit_columns(worksheet, min_width: int = 12) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = max(min_width, min(max_length + 2, 50))


async def export_journal_excel(
    db: AsyncSession,
    *,
    course_id: uuid.UUID,
    from_date: date,
    to_date: date,
) -> tuple[io.BytesIO, str]:
    if from_date > to_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="from_date must be on or before to_date.",
        )

    await _ensure_course_exists(db, course_id)

    plans_result = await db.execute(
        select(DailyWorkPlan)
        .where(
            DailyWorkPlan.course_id == course_id,
            DailyWorkPlan.plan_date >= from_date,
            DailyWorkPlan.plan_date <= to_date,
        )
        .options(
            selectinload(DailyWorkPlan.zone_tasks),
            selectinload(DailyWorkPlan.attendance).selectinload(DailyWorkerAttendance.worker),
        )
        .order_by(DailyWorkPlan.plan_date.asc())
    )
    plans = plans_result.scalars().all()

    issues_result = await db.execute(
        select(Issue)
        .where(
            Issue.course_id == course_id,
            cast(Issue.created_at, SaDate) >= from_date,
            cast(Issue.created_at, SaDate) <= to_date,
        )
        .options(selectinload(Issue.reporter))
        .order_by(Issue.created_at.asc())
    )
    issues = issues_result.scalars().all()

    all_worker_ids: set[uuid.UUID] = set()
    for plan in plans:
        for task in plan.zone_tasks:
            all_worker_ids.update(_parse_worker_ids(task.assigned_worker_ids))
    users_map = await _load_users_map(db, list(all_worker_ids))

    workbook = Workbook()

    work_sheet = workbook.active
    work_sheet.title = "작업일지"
    work_headers = [
        "날짜",
        "날씨",
        "기온",
        "구역",
        "작업내용",
        "예초높이(mm)",
        "담당자",
        "완료시간",
        "상태",
        "비고",
    ]
    _style_header_row(work_sheet, work_headers)
    work_row = 2
    for plan in plans:
        temp_range = ""
        if plan.temperature_min is not None or plan.temperature_max is not None:
            temp_range = f"{plan.temperature_min or ''}~{plan.temperature_max or ''}"
        if not plan.zone_tasks:
            work_sheet.append(
                [
                    plan.plan_date.isoformat(),
                    plan.weather,
                    temp_range,
                    "",
                    "",
                    "",
                    "",
                    "",
                    plan.status,
                    plan.special_notes or "",
                ]
            )
            work_row += 1
            continue
        for task in plan.zone_tasks:
            names = [
                users_map[wid].name
                for wid in _parse_worker_ids(task.assigned_worker_ids)
                if wid in users_map
            ]
            completed_at = (
                task.completed_at.astimezone(timezone.utc).strftime("%H:%M")
                if task.completed_at
                else ""
            )
            work_sheet.append(
                [
                    plan.plan_date.isoformat(),
                    plan.weather,
                    temp_range,
                    task.zone,
                    ", ".join(task.task_types),
                    task.mowing_height_mm,
                    ", ".join(names),
                    completed_at,
                    task.status,
                    task.notes or plan.special_notes or "",
                ]
            )
            work_row += 1
    _auto_fit_columns(work_sheet)

    attendance_sheet = workbook.create_sheet("근태현황")
    attendance_headers = [
        "날짜",
        "성명",
        "출근상태",
        "출근시간",
        "퇴근시간",
        "근무시간(h)",
    ]
    _style_header_row(attendance_sheet, attendance_headers)
    for plan in plans:
        for record in plan.attendance:
            worker_name = record.worker.name if record.worker else "Unknown"
            attendance_sheet.append(
                [
                    plan.plan_date.isoformat(),
                    worker_name,
                    record.status,
                    record.start_time or "",
                    record.end_time or "",
                    record.working_hours if record.working_hours is not None else "",
                ]
            )
    _auto_fit_columns(attendance_sheet)

    issues_sheet = workbook.create_sheet("이슈현황")
    issues_headers = [
        "날짜",
        "유형",
        "우선순위",
        "제목",
        "신고자",
        "상태",
        "해결일시",
    ]
    _style_header_row(issues_sheet, issues_headers)
    for issue in issues:
        reporter_name = issue.reporter.name if issue.reporter else "Unknown"
        resolved_str = (
            issue.resolved_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M")
            if issue.resolved_at
            else ""
        )
        issues_sheet.append(
            [
                issue.created_at.date().isoformat(),
                issue.issue_type,
                issue.priority,
                issue.title,
                reporter_name,
                issue.status,
                resolved_str,
            ]
        )
    _auto_fit_columns(issues_sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"GreenCare_작업일지_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"
    return buffer, filename
